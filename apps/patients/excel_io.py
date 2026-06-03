"""Экспорт/импорт пациентов в Excel (openpyxl)."""
from datetime import datetime, date
from django.http import HttpResponse


def _parse_date(v):
    if not v:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def export_patients_xlsx():
    from openpyxl import Workbook
    from .models import Patient
    wb = Workbook()
    ws = wb.active
    ws.title = "Пациенты"
    headers = ["Фамилия", "Имя", "Отчество", "Дата рождения", "Пол", "Телефон", "Доп. телефон", "Адрес", "Заметки"]
    ws.append(headers)
    for p in Patient.objects.all().order_by("last_name", "first_name"):
        ws.append([
            p.last_name, p.first_name, p.middle_name or "",
            p.birth_date.strftime("%d.%m.%Y") if p.birth_date else "",
            p.get_gender_display() if p.gender else "",
            p.phone, p.phone2 or "", p.address or "", p.notes or "",
        ])
    for i, w in enumerate([18, 16, 16, 16, 10, 18, 16, 30, 30], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="patients.xlsx"'
    wb.save(resp)
    return resp


def import_patients_xlsx(file_obj):
    """Колонки: Фамилия, Имя, [Отчество], [Дата рождения], [Пол], Телефон, [Доп.тел], [Адрес], [Заметки].
    Дедуп по (Фамилия+Имя+Телефон). Возвращает (created, updated, errors)."""
    from openpyxl import load_workbook
    from .models import Patient
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    created = updated = 0
    errors = []
    rows = ws.iter_rows(values_only=True)
    next(rows, None)
    GENDER = {"мужской": "male", "женский": "female", "м": "male", "ж": "female", "male": "male", "female": "female"}
    for i, row in enumerate(rows, start=2):
        if not row or all(c in (None, "") for c in row):
            continue
        g = lambda idx: (str(row[idx]).strip() if len(row) > idx and row[idx] not in (None, "") else "")
        last, first = g(0), g(1)
        if not (last or first):
            errors.append(f"Строка {i}: нет имени/фамилии")
            continue
        phone = g(5)
        defaults = {
            "middle_name": g(2),
            "birth_date": _parse_date(row[3] if len(row) > 3 else None),
            "gender": GENDER.get(g(4).lower(), ""),
            "phone2": g(6), "address": g(7), "notes": g(8),
        }
        obj, was_created = Patient.objects.update_or_create(
            last_name=last, first_name=first, phone=phone, defaults=defaults
        )
        created += 1 if was_created else 0
        updated += 0 if was_created else 1
    return created, updated, errors

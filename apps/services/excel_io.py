"""Экспорт/импорт прейскуранта услуг в Excel (openpyxl)."""
from decimal import Decimal, InvalidOperation
from django.http import HttpResponse


def _num(v, default=0):
    if v in (None, ""):
        return Decimal(default)
    try:
        return Decimal(str(v).replace(",", ".").replace(" ", ""))
    except (InvalidOperation, ValueError):
        return Decimal(default)


def export_services_xlsx():
    from openpyxl import Workbook
    from .models import Service
    wb = Workbook()
    ws = wb.active
    ws.title = "Прейскурант"
    headers = ["Категория", "Код", "Услуга", "Цена", "Цена ДМС", "Длительность (мин)", "Активна"]
    ws.append(headers)
    for s in Service.objects.select_related("category").order_by("category__name", "name"):
        ws.append([
            s.category.name if s.category else "",
            s.code or "",
            s.name,
            float(s.price),
            float(s.dms_price) if s.dms_price is not None else "",
            s.duration,
            "да" if s.is_active else "нет",
        ])
    for i, w in enumerate([26, 14, 40, 12, 12, 18, 10], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="services.xlsx"'
    wb.save(resp)
    return resp


def import_services_xlsx(file_obj):
    """Импорт: колонки Категория, Код, Услуга, Цена, [Цена ДМС], [Длительность], [Активна].
    Возвращает (created, updated, errors)."""
    from openpyxl import load_workbook
    from .models import Service, ServiceCategory
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    created = updated = 0
    errors = []
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # header
    for i, row in enumerate(rows, start=2):
        if not row or all(c in (None, "") for c in row):
            continue
        cat_name = (row[0] or "").strip() if len(row) > 0 and row[0] else ""
        code = (str(row[1]).strip() if len(row) > 1 and row[1] else "")
        name = (str(row[2]).strip() if len(row) > 2 and row[2] else "")
        if not name:
            errors.append(f"Строка {i}: нет названия услуги")
            continue
        price = _num(row[3] if len(row) > 3 else 0)
        dms = row[4] if len(row) > 4 else None
        dur = int(_num(row[5], 30)) if len(row) > 5 and row[5] not in (None, "") else 30
        active_raw = (str(row[6]).strip().lower() if len(row) > 6 and row[6] is not None else "да")
        is_active = active_raw not in ("нет", "no", "0", "false")

        cat = None
        if cat_name:
            cat, _ = ServiceCategory.objects.get_or_create(name=cat_name)
        defaults = {"category": cat, "price": price, "duration": dur, "is_active": is_active, "code": code}
        if dms not in (None, ""):
            defaults["dms_price"] = _num(dms)
        obj, was_created = Service.objects.update_or_create(name=name, defaults=defaults)
        created += 1 if was_created else 0
        updated += 0 if was_created else 1
    return created, updated, errors
